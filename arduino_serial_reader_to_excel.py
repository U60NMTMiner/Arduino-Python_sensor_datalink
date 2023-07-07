import serial
import pandas as pd
import datetime
import matplotlib.pyplot as plt      # Importing Libraries
import sys
import os
import openpyxl as xl

now = datetime.datetime.now()        # Setting up datetime to create unique file name for each test
filename = now.strftime("%Y-%m-%d_%H-%M") + "_data.xlsx"

ser = serial.Serial('COM10', 9600)    # Setup for reading the serial communication
print("Serial port opened")
print("For best results, please allow smoke sensors to warm to operating temperature for 30 minutes before use!")
print()

prelim = ser.readline().decode().strip()  # Read just the first line of serial for sensor names
prelim = prelim.split(' ')                # Put the sensor ID's sent by the Arduino into a list
senslen = len(prelim)                     # Figure out the number of active sensors
for i in range(senslen):                  # Pull out the sensor IDs from the list
    exec(f"{prelim[i]} = {0}")
del i                                     # Clean up unneeded variable
print("Sensor names imported")

sensloc = ser.readline().decode().strip()  # Read just the second line of serial for sensor locations
sensloc = sensloc.split(' ')
print("Sensor locations imported")

n = 0                                     # Start master index at zero

data_df = pd.DataFrame(columns=[prelim])                          # Start up dataframe
data_df.loc[0] = [0] * len(prelim)                                # Temporarily fill with zeros
row_data = [0] * len(prelim)                                      # Create list to contain each second's data
combined_df = pd.DataFrame(columns=[prelim])                      # Start up master dataframe to hold all data
empty_df = data_df                                                # Dataframe of just zeros


def format_file():                                                # Modify the output .xlsx file to work with Mine_Evacuation.py
    data_workbook = xl.load_workbook(filename)                    # Load the raw data file
    wsedit = data_workbook['Sheet2']
    for q, value in enumerate(sensloc):
        cell = wsedit.cell(row=4, column=3 + q)                   # Assign node location to each sensor
        cell.value = value
    del q

    data_workbook.save(filename + "mod.xlsx")                     # Save new file
    os.remove("C:\\Users\\Sean\\PycharmProjects\\Gas sensor\\" + filename)  #clean up preliminary file

    source_workbook = xl.load_workbook("M_E_DataTemplate.xlsx")   # Load the template data file
    dupe_workbook = xl.Workbook()                                 # Create new spreadsheet to preserve original
    for sheet_name in source_workbook.sheetnames:                 # For each sheet in the template...
        source_sheet = source_workbook[sheet_name]                # Get the source sheet...
        new_sheet = dupe_workbook.create_sheet(title=sheet_name)  # Create a new sheet in the new workbook...
        for row in source_sheet.iter_rows(values_only=True):      # And copy the values from the source sheet to the new sheet
            new_sheet.append(row)
        del row
    del sheet_name
    del dupe_workbook['Sheet']                                    # Get rid of copy operation artefact
    source_workbook.close()                                       # Close the original without saving to prevent accidental overwrites

    datalen = len(combined_df.index)                              # Determine how much data was collected
    ws2edit = dupe_workbook['Sheet2']                             # Prepare the new workbook to make edits
    ws2edit.delete_rows(5, 13)                                    # Clear out the example data
    for row in range(5, 1 + 4 + datalen):
        for column in range(2, 199 + 1):
            cell = ws2edit.cell(row=row, column=column)
            cell.value = 0                                        # Put in zeros to fill any would-be empty cells
    for w in range(datalen):
        ws2edit.cell(row=w + 5, column=1).value = w               # Put in the new index
    del w
    del column
    del row

    # Based on the sensors' node locations, copy data into the new workbook

    # Start with gas sensors
    gas_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]  # Telling the program which columns in the data spreadsheet to read for the gas sensors
    o = 0
    for item in sensloc[:16]:       # We know that only the first 16 sensors are gas sensors (only 16 analog ports on the Arduino Mega)
        ref_cell = str(item)        # Sensor location we are looking for
        for col in ws2edit.iter_cols(min_row=3, max_row=3, min_col=167, max_col=199 + 1):
            search_cell = ws2edit.cell(row=3, column=col[0].col_idx).value    # Where the program is currently looking
            if str(search_cell) == ref_cell and search_cell != "Fire":        # If it finds something...
                print("Matching sensor found at node " + str(search_cell))  #eventually, this needs to copy the data from wsedit to ws2edit
                for row in wsedit.iter_rows(min_row=5, max_row=datalen + 4, min_col=col[0].col_idx, max_col=col[0].col_idx):  # Look at all the data belonging to the sensor
                    ws2edit.cell(row[0].row, col[0].col_idx).value = wsedit.cell(row[0].row, gas_cols[o]).value  # And copy it over
                o += 1
    del o

    # Next airspeed sensors
    #airspeed_cols = [19, 20, 21, 22, 23, 24, 25, 26]  # Telling the program which columns in the data spreadsheet to read for the gas sensors
    airspeed_cols = [19]  # Until the TCA9548A multiplexers get here, only one sensor allowed on the I2C bus :(
    p = 0
    #for item in sensloc[16:23]:  #Until the multiplexers get here, just one sensor
    for item in sensloc[17-1]:
        ref_cell = str(item)        # Sensor location we are looking for
        for col in ws2edit.iter_cols(min_row=3, max_row=3, min_col=35, max_col=66 + 1):
            search_cell = ws2edit.cell(row=3, column=col[0].col_idx).value                  # Reading all the sensor numbers
            if str(search_cell) == ref_cell and search_cell != "Fire":                      # If it finds a match...
                print("Matching sensor found at node " + str(search_cell))                  #
                for row in wsedit.iter_rows(min_row=5, max_row=datalen + 4, min_col=col[0].col_idx, max_col=col[0].col_idx):  # Look at all the data belonging to the sensor
                    ws2edit.cell(row[0].row, col[0].col_idx).value = wsedit.cell(row[0].row, airspeed_cols[p]).value          # And copy it over
                p += 1
    del p

    dupe_workbook.save("duplicate.xlsx")                          # Save the new workbook (Will overwrite itself each time the program is run)
    dupe_workbook.close()                                         # Close all workbooks
    data_workbook.close()
    os.remove("C:\\Users\\Sean\\PycharmProjects\\Gas sensor\\" + filename + "mod.xlsx")  #clean up previous stage's file
    return


while True:                                    # Main program loop
    data = ser.readline().decode().strip()     # Continuously pull data from the serial stream
    if data:                                   # Make sure there is something to read
        if "!" in data:                        # Listen for the exit condition from the Arduino
            with pd.ExcelWriter(filename) as writer:
                empty_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
                combined_df.to_excel(writer, sheet_name='Sheet2', index=True, header=True, startrow=2, index_label='Timestamp', startcol=1)  # If the condition is met, save the data...
                empty_df.to_excel(writer, sheet_name='Sheet3', index=False, header=False)
                empty_df.to_excel(writer, sheet_name='Sheet4', index=False, header=False)
                print("New spreadsheet generating...")
                print()
            format_file()
            print()
            print("Success!")
            print("Formatted data exported as file: " + filename + "mod.xlsx")  # Then export the file
            sys.exit(0)                                  # Then shut down this program

        data = data.split(',')                 # Separate the sensors' data
        del data[-1]                           # Account for the extra comma at the end

        row_data = [float(item[3:]) for item in data]    #put the data into a list, everything from the 8th character on
        combined_df = pd.concat([combined_df, pd.DataFrame([row_data], columns=[prelim])], ignore_index=True)

        plt.figure(1)                          # Set up the data plot
        plt.ion()                              # Enable interactive graph (plot will automatically redraw)
        plt.ylim([-5, 10000])                    # Set y-axis bounds
        plt.xlim([n - 500, n + 5])             # Set scrolling x-axis bounds
        plt.plot(combined_df['S01'].iloc[-500:], 'r-', label="Gas Sensor 1")
        plt.plot(combined_df['S02'].iloc[-500:], 'b-', label="Gas Sensor 2")  # Set gas sensor data colors, and
        plt.plot(combined_df['A01'].iloc[-500:]*100, 'g-', label="Airspeed Sensor 1")  # Plot the last 1000 datapoints of each
        plt.title("Sensor Datafeed [2sec delay]")   # Give the user more context
        plt.xlabel("Time (seconds)")
        plt.ylabel("Hazardous Gas Concentration (ppm)")
        if n == 0:                                  # Only add one legend, not a new one for each loop
            plt.legend(loc="upper left")
        plt.pause(0.0001)                           # Mandatory pause command
        n += 1                                      # Move master index forward
