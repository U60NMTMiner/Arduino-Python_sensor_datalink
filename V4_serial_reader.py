import datetime
import time
import serial
import sys
import select
import os
import openpyxl as xl
import Modules.functions as func

start_time = time.time()

# Load and initialize the configuration
config = func.open_file("config")

# Get working directory from config. If it is an empty string, set cwd to the location of this file.
cwd = config['Working_Directory']
if not cwd:
    cwd = func.default_path()

# Once the sensors have been installed and wired into the multiplexer arrays, coordinate locations will be assigned here
# Sensor layout and coordinate system are defined here:
# https://docs.google.com/presentation/d/1iU9-4sDkxN5r5blLVAVAElmQhg7pdddcT2UxLzKhD1Q/edit?usp=sharing
Airspeed_Sensor_Coordinates = config['Airspeed_Sensor_Coordinates']
Gas_Sensor_Coordinates = config['Gas_Sensor_Coordinates']
Temperature_Sensor_Coordinates = config['Temperature_Sensor_Coordinates']

"""
Note: I disabled write-only mode.
Reasons: In either case, its only temporarily useful.
        When using Mode 1   -> You're creating a new file, there's nothing to read anyways
        When using Mode 2   -> If you save and then load the file back up, it's now in read/write mode anyways
        In both cases       -> We write to file after 3 loops of the buffer. We're not dumping large amounts of data
"""
# Create the spreadsheet
filename = f'{datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")}_data.xlsx'  # Set up unique filename for exported data
mainWB = xl.Workbook()  # Create spreadsheet to store data
datasheet = mainWB.create_sheet("Data")  # Name the sheet we want to work on


# del mainWB["sheet"]  # Delete artifact (only need it when not using write_only mode)


# Convert bytes to integers
def bytes_to_int(in_bytes):  # Reconstructs 4 bytes of data into a long int
    value = (in_bytes[0] << 24) | (in_bytes[1] << 16) | (in_bytes[2] << 8) | in_bytes[3]
    return value


def bytes_to_int2(in_bytes):
    """
    This is literally the same as the previous function, but just scales to size.
    """
    value = 0
    size = len(in_bytes)
    for i in range(0, size):
        bit_shift = 8 * (size - i - 1)
        value = value | (in_bytes[i] << bit_shift)
    return value


def inputTimeout(prompt, timeout):
    prompt_time = time.time()
    while time.time() - prompt_time < timeout:
        print(prompt, end='', flush=True)
        if sys.stdin in select.select([sys.stdin], [], [], timeout)[0]:
            return sys.stdin.readline().strip()
    return None


def waitUntilReady(serial_connection):
    """
    Waits until the Arduino sends the ready message
    :param serial_connection: Connection to wait on.
    """
    while True:
        # For the 'mirroring' before actual data is coming in, use built-in decoding
        intro = serial_connection.readline().decode().strip()
        print(intro)
        if intro == "Nano Pass-Through ready":
            break


def getCurrentTime() -> float:
    """
    :return: Current time relative to start time
    """
    return time.time() - start_time


def keepConstantTiming(last_time, interval):
    """
    Ensures a constant timing between data dumps
    :param last_time: Time since last break ended
    :param interval: Time to wait from last break until starting next loop
    """
    time_passed = getCurrentTime() - last_time
    time_remaining = interval - time_passed
    wait_time = time_remaining / 4

    # Run this loop 3 times to get 75% of the way done with the wait time. (Large jumping)
    for i in range(0, 3):
        print('.', end='')
        time.sleep(wait_time)

    # Short jumping to get a more precise end time.
    while getCurrentTime() - last_time < interval:
        pass


def saveWorkbook():
    """
    Saves the workbook
    """
    mainWB.save(f'{cwd}/{filename}')


def readData(serialConnection, sessionData, refinedData):
    """
    Reads the data from the serial connection and stores the data into the dicts
    :param serialConnection: Connection to the Arduino
    :param sessionData: Data related to the current session
    :param refinedData: Data refined from previous buffer loops
    """

    # Python doesn't have pass-by-reference for primitive data types, so we're using a dict() to pass everything
    buffer = sessionData['buffer']
    last_time = sessionData['last_time']
    refinedAData, refinedTData, refinedSData = refinedData['a'], ['s'], ['t']

    data = serialConnection.read()  # As data comes in, do not apply decoding since everything is meant to be in binary
    buffer += data  # Put it into a buffer for short-term storage

    BadData = False

    # Check if the Arduino returned an error
    if buffer.endswith(b"?????"):
        print("\033[93m" + "Warning: Nano pass-thru returned an error" + "\033[0m")
        buffer = b""  # If so, make sure to reset the buffer
        time.sleep(1)

    # Check for terminating symbol
    if buffer.endswith(b"~~~~~"):  # And if the terminating symbol comes in...
        SplitData = list(buffer)  # Split up the byte array into list items that are easier to work with

        AirFlowData = refinedData['AirFlow']

        # Check if Airflow data
        if SplitData[0] == 65 and len(
                SplitData) == 120:  # Check if it starts with 0x65 (DEC for "A" for airspeed) and make sure all of it is there
            cleanAData = [item for index, item in enumerate(SplitData) if
                          (index + 1) % 5 != 1]  # Clean out the sensor type identifiers
            cleanAData = cleanAData[:-4]  # Remove the last remaining bit of the 5-character terminator symbol
            for i in range(0, len(cleanAData), 4):
                convertedChunk = bytes_to_int(
                    cleanAData[i:i + 4])  # Convert the data from binary to integers, in blocks of 4 bytes
                v = convertedChunk / 10000
                refinedAData.append(v)  # Convert integer to float using the opposite operation as the Arduino Mega made
                print(convertedChunk / 10000)
                AirFlowData.append(v * 0.00112903)
            # print(refinedAData)
            refinedData['a'].append(refinedAData)
            refinedData['AirFlow'].append(AirFlowData)
            # AirFlowData = [airspeed * 0.00112903 for airspeed in refinedAData]  # Spin off a new list for airflow by multiplying by cross-sectional area

        # Check if Smoke data
        elif SplitData[0] == 83 and len(
                SplitData) == 470:  # Check for DEC "S" for smoke data and make sure all of it is there
            cleanSData = [item for index, item in enumerate(SplitData) if
                          (index + 1) % 5 != 1]  # Clean out the sensor type identifiers
            cleanSData = cleanSData[:-4]  # Remove the last of the 5-character terminator symbol
            for i in range(0, len(cleanSData), 4):
                convertedChunk = bytes_to_int(
                    cleanSData[i:i + 4])  # Convert the data from binary to integers, in blocks of 4 bytes
                refinedSData.append(convertedChunk)  # No need to convert to float here, it is intended to be an integer
            print(refinedSData)
            refinedData['s'].append(refinedSData)

        # Check if Temperature data
        elif SplitData[0] == 84 and len(
                SplitData) == 450:  # Check for DEC "T" for temperature data and make sure all of it is there
            cleanTData = [item for index, item in enumerate(SplitData) if
                          (index + 1) % 5 != 1]  # Clean out the sensor type identifiers
            cleanTData = cleanTData[:-4]  # Remove the last of the 5-character terminator symbol
            for i in range(0, len(cleanTData), 4):
                convertedChunk = bytes_to_int(
                    cleanTData[i:i + 4])  # Convert the data from binary to integers, in blocks of 4 bytes
                refinedTData.append(
                    convertedChunk / 10000)  # Convert integer to float using the opposite operation as the Arduino Mega made
            print(refinedTData)
            refinedData['t'].append(refinedTData)

        else:
            print("\033[93m" + "Warning: Bad serial TX/RX. Suspected bad data set will be dumped." + "\033[0m")
            BadData = True  # If something isn't quite right about the data, don't let it get saved to the spreadsheet later

        buffer = b""  # Purge buffer for next set of incoming data

        # Check if 3 sets of data have arrived since last write.
        x = (sessionData['index'] + 1) % 3
        print(f'x:{x}\nindex:{sessionData["index"]}')
        if x == 0:  # And after all 3 sets of data have arrived...
            current_time = getCurrentTime()  # Check the time
            print("Timestamp: " + str(int(current_time)) + " seconds\n")  # Print timestamp for user

            if not sessionData['Header']:  # If the spreadsheet was not already given a header...
                headerList = ([str(filename)]
                              + ["Airflow (m^3/s)" for _ in refinedData['AirFlow']]
                              + ["Air Velocity (m/s)" for _ in refinedData['a']]
                              + ["Temperature (C)" for _ in refinedData['t']]
                              + ["Gas (ppm)" for _ in refinedData['s']])
                datasheet.append(headerList)  # Create one based on the amount of data being received
                headerList = (["Time (seconds)"]
                              + Airspeed_Sensor_Coordinates
                              + Airspeed_Sensor_Coordinates
                              + Temperature_Sensor_Coordinates
                              + Gas_Sensor_Coordinates)
                datasheet.append(headerList)
                del headerList  # Clean up memory
                sessionData['Header'] = True  # Now it has a header

            if not BadData:  # If the data was acceptable, add it to the spreadsheet
                ExportList = ([int(current_time)]
                              + refinedData['AirFlow']
                              + refinedData['a']
                              + refinedData['t']
                              + refinedData['s'])

                print(f'Refined Data A: {refinedData["a"]}\n'
                      f'Refined Data T: {refinedData["t"]}\n'
                      f'Refined Data S: {refinedData["s"]}\n'
                      f'ExportList: {ExportList}')
                datasheet.append(ExportList)
            elif BadData:
                print("\033[93m" + "Suspected bad data set was dumped." + "\033[0m")

            refinedData['a'], refinedData['s'], refinedData['t'] = [], [], []  # Make room for new data decoding

            saveWorkbook()
            # datasheet = mainWB["Data"]
            print("Spreadsheet updated.")

            keepConstantTiming(last_time, 10)

            last_time = current_time
            print()
            print("Requesting new data...")
            serialConnection.write(b"\x47")  # Send the "request data" command to the Nano

    # Update the session data with what we've read.
    sessionData['buffer'] = buffer
    sessionData['last_time'] = last_time
    return


def main():
    # Initializing variables during startup
    sessionData = dict()
    sessionData['buffer'] = b""
    sessionData['Header'] = False
    sessionData['index'] = 0
    refinedData = dict()
    refinedData['a'], refinedData['s'], refinedData['t'] = [], [], []
    refinedData['AirFlow'] = []

    # Open serial port
    ser = serial.Serial(config['Serial_Port'], 1000000)
    print("\033[94m" + "Serial connection opened" + "\033[0m")
    waitUntilReady(ser)
    ser.write(b"\x47")  # Request first data set

    # IDK why, but for some reason this was in the original code, so we keep it :D
    current_time = getCurrentTime()
    print(f"Current time: {current_time}\t"
          f"Startup time: {start_time}\t"
          f"Time Elapsed during Startup: {current_time - start_time}")
    sessionData['last_time'] = current_time

    try:
        # Begin the infinite data reading!
        while True:
            readData(ser, sessionData, refinedData)
            sessionData['index'] += 1

    except KeyboardInterrupt:  # Gracefully save data and shut down the program instead of crashing
        print("\033[93m" + "[Keyboard interrupt]" + "\033[0m")
        if ser.is_open:
            ser.close()  # Close the serial connection
            print("\033[94m" + "Serial connection closed" + "\033[0m")

        # try:
        print("Saving spreadsheet...")
        saveWorkbook()  # Save the spreadsheet
        mainWB.close()
        print(f"Spreadsheet saved: \033[32m{cwd}/{filename}\033[0m")
        # except FileNotFoundError:  # ... Unless it was already saved
        #     print(f"Spreadsheet was already saved: \033[32m{cwd}/{filename}\033[0m")

        final = input("Enter anything to exit program, or print 'del' to delete the spreadsheet:")
        if final == "del":
            os.remove(f'{cwd}/{filename}')
            print("File deleted.")
        else:
            print("Exiting program.")

    sys.exit(0)  # Show "success" exit code


if __name__ == "__main__":
    main()
